#!/usr/bin/env python3
"""Build a filled GTM workbook from an answers JSON file.

    python build_gtm_workbook.py answers.json -o GTM_filled.xlsx
    python build_gtm_workbook.py --schema

Every top-level key is optional. Omitted sections produce a blank tab with the
headers in place, so a half-answered JSON yields a half-filled workbook that the
team can finish by hand. Missing values are written as "TBD" rather than guessed.
"""

import argparse
import json
import sys

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    HAVE_OPENPYXL = True
except ImportError:  # --schema still works without openpyxl
    HAVE_OPENPYXL = False

FONT = "Arial"
INK = "1A1A1A"
ACCENT = "1F3864"
BAND = "DDEBF7"
TBD = "TBD"

if HAVE_OPENPYXL:
    H1 = Font(name=FONT, size=14, bold=True, color="FFFFFF")
    H2 = Font(name=FONT, size=11, bold=True, color=ACCENT)
    LABEL = Font(name=FONT, size=10, bold=True, color=INK)
    BODY = Font(name=FONT, size=10, color=INK)
    MUTED = Font(name=FONT, size=10, italic=True, color="808080")
    INPUT_FONT = Font(name=FONT, size=10, color="0000FF")  # blue = hardcoded input
    TITLE_FILL = PatternFill("solid", start_color=ACCENT)
    BAND_FILL = PatternFill("solid", start_color=BAND)
    WRAP = Alignment(wrap_text=True, vertical="top")


def title(ws, text, span=6):
    ws["A1"] = text
    ws["A1"].font = H1
    for c in range(1, span + 1):
        ws.cell(row=1, column=c).fill = TITLE_FILL
    ws.row_dimensions[1].height = 22


def section(ws, row, text, span=6):
    ws.cell(row=row, column=1, value=text).font = H2
    for c in range(1, span + 1):
        ws.cell(row=row, column=c).fill = BAND_FILL
    return row + 1


def pair(ws, row, label, value, col=1):
    ws.cell(row=row, column=col, value=label).font = LABEL
    cell = ws.cell(row=row, column=col + 1, value=value if value else TBD)
    cell.font = BODY if value else MUTED
    cell.alignment = WRAP
    return row + 1


def table(ws, row, headers, rows, widths=None):
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = LABEL
        c.fill = BAND_FILL
        c.alignment = WRAP
    for i, w in enumerate(widths or [22] * len(headers), start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    row += 1
    for r in rows:
        for i, v in enumerate(r, start=1):
            c = ws.cell(row=row, column=i, value=v if v not in (None, "") else TBD)
            c.font = BODY if v not in (None, "") else MUTED
            c.alignment = WRAP
        row += 1
    return row



# --- tabs -------------------------------------------------------------------


def tab_brainstorm(wb, data):
    ws = wb.create_sheet("Opportunity Brainstorm")
    title(ws, "OPPORTUNITY BRAINSTORM", span=6)
    ideas = data.get("opportunity_ideas", []) or [{} for _ in range(5)]
    rows = [[
        i.get("id", f"Idea {n}"), i.get("industry_segment", ""), i.get("problem_statement", ""),
        i.get("target_audience", ""), i.get("technology_opportunity", ""), i.get("name", ""),
    ] for n, i in enumerate(ideas, start=1)]
    r = table(ws, 3,
              ["Idea", "Industry and Market Segment", "Problem Statement (Statement of Need)",
               "Target customers / audience", "Opportunity for Blockchain", "Name of the idea"],
              rows, widths=[10, 28, 60, 30, 55, 26])
    ws.cell(row=r + 1, column=1,
            value="Problem statements contain no solution. If no multi-party trust, verifiable "
                  "history, programmable settlement, credible commitment, or portable identity "
                  "requirement applies, write \"a database would do this\" and keep the row.").font = MUTED


def tab_decision_matrix(wb, data):
    ws = wb.create_sheet("Decision Matrix")
    dm = data.get("decision_matrix", {}) or {}
    factors = dm.get("factors", []) or [{"name": f"Factor {i}"} for i in range(1, 6)]
    ideas = dm.get("ideas", []) or [{"name": f"Name of idea {i}"} for i in range(1, 6)]
    scale = dm.get("scale", "1-5")

    title(ws, f"DECISION MATRIX (raw scores on a {scale} scale; weights should sum to 100)",
          span=2 * len(factors) + 3)
    ws.column_dimensions["A"].width = 26

    WEIGHT_ROW, FIRST = 5, 6
    raw_cols = [2 + 2 * i for i in range(len(factors))]
    total_col = raw_cols[-1] + 2
    rank_col = total_col + 1

    for i, (f, rc) in enumerate(zip(factors, raw_cols), start=1):
        ws.cell(row=3, column=rc, value=f"Factor {i}:").font = LABEL
        c = ws.cell(row=4, column=rc, value=f.get("name", f"Identify Factor {i}"))
        c.font = LABEL
        c.fill = BAND_FILL
        c.alignment = WRAP
        ws.cell(row=4, column=rc + 1, value="Weighted").font = MUTED
        ws.column_dimensions[get_column_letter(rc)].width = 18
        ws.column_dimensions[get_column_letter(rc + 1)].width = 11
    for col, label in [(total_col, "TOTAL"), (rank_col, "RANK")]:
        c = ws.cell(row=4, column=col, value=label)
        c.font = LABEL
        c.fill = BAND_FILL
        ws.column_dimensions[get_column_letter(col)].width = 10

    ws.cell(row=WEIGHT_ROW, column=1, value="Factor Weighting").font = LABEL
    for f, rc in zip(factors, raw_cols):
        c = ws.cell(row=WEIGHT_ROW, column=rc, value=f.get("weight", 0))
        c.font = INPUT_FONT  # blue: hardcoded input the user will change
    ws.cell(row=WEIGHT_ROW, column=total_col,
            value="=SUM(" + ",".join(f"{get_column_letter(c)}{WEIGHT_ROW}" for c in raw_cols) + ")").font = BODY
    ws.cell(row=WEIGHT_ROW, column=rank_col, value="<- must be 100").font = MUTED

    last = FIRST + len(ideas) - 1
    for n, idea in enumerate(ideas):
        row = FIRST + n
        ws.cell(row=row, column=1, value=idea.get("name", f"Name of idea {n + 1}")).font = BODY
        scores = idea.get("scores", [])
        for j, rc in enumerate(raw_cols):
            rl = get_column_letter(rc)
            v = scores[j] if j < len(scores) else 0
            ws.cell(row=row, column=rc, value=v).font = INPUT_FONT
            # anchor the weight row: the shipped template omits the $ and mis-scores Factor 1
            ws.cell(row=row, column=rc + 1, value=f"={rl}{row}*{rl}${WEIGHT_ROW}").font = BODY
        ws.cell(row=row, column=total_col,
                value="=SUM(" + ",".join(f"{get_column_letter(c + 1)}{row}" for c in raw_cols) + ")").font = LABEL
        tl = get_column_letter(total_col)
        ws.cell(row=row, column=rank_col,
                value=f"=RANK({tl}{row},{tl}${FIRST}:{tl}${last})").font = BODY

    note = ("Flip test: change the weight you are least sure of by +/-10. "
            "If the winner changes, the result rests on that judgment. "
            "If the top two totals are within 5%, the matrix has not decided anything.")
    ws.cell(row=last + 2, column=1, value=note).font = MUTED


def tab_mission(wb, data):
    ws = wb.create_sheet("Your Company Mission Statement")
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 90
    c = data.get("company", {})
    title(ws, f"COMPANY PROFILE: {c.get('name', '')}".strip(": "), span=2)
    r = 3
    for label, key in [
        ("Vision:", "vision"),
        ("Mission:", "mission"),
        ("Values:", "values"),
        ("Culture:", "culture"),
        ("Corporate Objective:", "corporate_objective"),
        ("Your Mission Critical Priority:", "mission_critical_priority"),
    ]:
        v = c.get(key, "")
        r = pair(ws, r, label, "; ".join(v) if isinstance(v, list) else v)
    r += 1
    r = section(ws, r, "Milestones", span=2)
    for m in c.get("milestones", []) or [""]:
        r = pair(ws, r, "", m)


def tab_icp(wb, data):
    ws = wb.create_sheet("B2B ICP Mapping")
    ws.column_dimensions["A"].width = 60
    ws.column_dimensions["B"].width = 90
    title(ws, "QUALIFYING QUESTIONS", span=2)
    o, m, b = data.get("offer", {}), data.get("target_market", {}), data.get("buyer", {})
    r = section(ws, 3, "Your Service / Product", span=2)
    for q, k in [
        ("1. Description of the problem your solution is solving", "problem"),
        ("2. Description of your solution", "solution"),
        ("3. Are there existing solutions that solve the same problem?", "existing_solutions"),
        ("   > 3a. If yes, how are you different?", "differentiation_vs_existing"),
        ("   > 3b. If no, why do you think that is?", "why_no_existing"),
        ("4. How is your solution different from what exists in the market?", "differentiation"),
    ]:
        r = pair(ws, r, q, o.get(k, ""))
    r += 1
    r = section(ws, r, "Your Target Market", span=2)
    for q, k in [
        ("1. How much revenue would they need for your solution to make sense?", "revenue_threshold"),
        ("2. How are they earning their revenue right now?", "revenue_model"),
        ("   > Does their business model make a purchase/partnership sensible?", "model_fit"),
        ("3. Why would they partner with / purchase your service?", "why_buy"),
        ("4. How much would they spend? OR what partnership model suits them?", "willingness_to_pay"),
        ("5. Can they afford your service / product?", "ability_to_pay"),
    ]:
        r = pair(ws, r, q, m.get(k, ""))
    r += 1
    r = section(ws, r, "Your Buyers' Persona", span=2)
    for q, k in [
        ("1. What pain points drove this persona to seek your solution?", "pain_points"),
        ("2. What motivates this persona to seek improvements?", "motivation"),
        ("3. What does your solution enable for this persona?", "enables"),
        ("4. What initiative leads them to seek your solution?", "initiative"),
        ("5. Are they Decision Makers?", "is_decision_maker"),
        ("   > 5a. If yes, who else must they convince?", "dm_influencers"),
        ("   > 5b. If no, who is the Decision Maker?", "decision_maker"),
        ("      > 5bi. What is the route to the Decision Maker?", "route_to_dm"),
        ("6. Are they Purse Holders (financial approver)?", "is_purse_holder"),
        ("   > 6a. If yes, what must they see to allocate budget?", "budget_requirements"),
        ("   > 6b. If no, who is the Purse Holder and what must they see?", "purse_holder"),
        ("      > 6bi. What is the route to Budget?", "route_to_budget"),
    ]:
        r = pair(ws, r, q, b.get(k, ""))


def tab_personas(wb, data):
    ws = wb.create_sheet("Personas Mapping")
    title(ws, "PERSONAS MAPPING", span=8)
    default = ["A - Must Win", "B - Tier 1", "C - Tier 2", "D - Tier 3"]
    personas = data.get("personas", []) or [{"tier": t} for t in default]
    headers = [
        "Persona", "Industry", "Revenue", "Employee Size", "Countr(ies)",
        "Job Title", "Work Experience", "Education Background", "Relevance",
        "Interaction Preferences", "Content Types Preferred", "Content Sources",
    ]
    rows = []
    for p in personas:
        co, pe, at = p.get("company", {}), p.get("persona", {}), p.get("attention", {})
        rows.append([
            p.get("tier", ""), co.get("industry", ""), co.get("revenue", ""),
            co.get("employee_size", ""), co.get("countries", ""),
            pe.get("job_title", ""), pe.get("work_experience", ""), pe.get("education", ""),
            p.get("relevance", ""), at.get("interaction_preferences", ""),
            at.get("content_types", ""), at.get("content_sources", ""),
        ])
    table(ws, 3, headers, rows, widths=[16, 20, 16, 14, 16, 22, 22, 20, 40, 30, 30, 30])


def tab_methodologies(wb, data):
    ws = wb.create_sheet("Sales Methodologies")
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 100
    title(ws, "B2B SALES METHODOLOGIES", span=2)
    m = data.get("methodologies", {})
    q = m.get("qualification", {})
    r = section(ws, 3, f"Qualification framework: {q.get('framework', TBD)}", span=2)
    for letter, meaning in (q.get("definitions", {}) or {"": ""}).items():
        r = pair(ws, r, letter, meaning)
    r += 1
    p = m.get("philosophy", {})
    r = section(ws, r, f"Selling philosophy: {p.get('name', TBD)}", span=2)
    r = pair(ws, r, "How we run it", p.get("notes", ""))
    r = pair(ws, r, "Stage gate it enforces", p.get("stage_gate", ""))


def tab_cycle(wb, data):
    ws = wb.create_sheet("Sales Cycle")
    title(ws, "B2B SALES CYCLE", span=6)
    default_stages = [
        "Stage 1: Discovery / Qualification Call",
        "Stage 2: Further Discovery / Alignment Call",
        "Stage 3: Alignment Call",
        "Stage 4: POC",
        "Stage 5: Objection Handling / Further Alignment / Price Check",
        "Stage 6: Initial Proposal / Objection Handling / Negotiation",
        "Stage 7: Final Proposal / Verbal Commit",
        "Stage 8: Won / Lost",
    ]
    stages = data.get("sales_cycle", []) or [{"stage": s} for s in default_stages]
    rows = [[
        s.get("stage", ""), s.get("entry_evidence", ""), s.get("goal", ""),
        s.get("exit_criteria", ""), s.get("conversion_target", ""), s.get("median_days_target", ""),
    ] for s in stages]
    table(ws, 3,
          ["Stage", "Entry evidence (buyer did)", "Goal", "Exit criteria (buyer did)",
           "Conv. to next stage", "Median days in stage"],
          rows, widths=[38, 40, 36, 46, 18, 18])


def tab_outreach(wb, data):
    ws = wb.create_sheet("Cold Outreach")
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 90
    title(ws, "B2B COLD OUTREACH", span=6)
    c = data.get("cold_outreach", {})
    e = c.get("cold_email", {})
    r = section(ws, 3, "Cold Email: BASHO / WYWYN (Why You, Why You Now)", span=6)
    r = pair(ws, r, "Trigger (Why You Now)", e.get("trigger", ""))
    r = pair(ws, r, "Relevance bridge (Why You)", e.get("bridge", ""))
    r = pair(ws, r, "Value hypothesis", e.get("value_hypothesis", ""))
    r = pair(ws, r, "Interest-based CTA", e.get("cta", ""))
    r = pair(ws, r, "Subject line", e.get("subject", ""))
    r = pair(ws, r, "Full draft", e.get("body", ""))
    r += 1
    r = section(ws, r, "Cold Call: 20-second elevator pitch", span=6)
    r = pair(ws, r, "Pitch", c.get("cold_call_pitch", ""))
    r = pair(ws, r, "Opening question", c.get("cold_call_question", ""))
    r += 1
    r = section(ws, r, "Touches", span=6)
    touches = c.get("touches", []) or [{} for _ in range(8)]
    rows = [[t.get("n", i + 1), t.get("day", ""), t.get("channel", ""), t.get("content", "")]
            for i, t in enumerate(touches)]
    table(ws, r, ["#", "Day", "Channel", "Content / angle"], rows, widths=[6, 8, 18, 70])


def tab_objections(wb, data):
    ws = wb.create_sheet("Objection Handling")
    title(ws, "OBJECTION HANDLING (Listen, Acknowledge, Explore, Respond)", span=5)
    rows = [[o.get("objection", ""), o.get("underlying_cause", ""), o.get("exploring_question", ""),
             o.get("response", ""), o.get("proof_asset", "")]
            for o in (data.get("objections", []) or [{}])]
    table(ws, 3,
          ["Objection heard", "Likely underlying cause", "Exploring question", "Response", "Proof asset"],
          rows, widths=[28, 34, 40, 44, 24])


def tab_partnerships(wb, data):
    ws = wb.create_sheet("Partnership Goals")
    title(ws, "PARTNERSHIP GOALS MATRIX", span=11)
    rows = [[
        p.get("no", i + 1), p.get("goal_type", ""), p.get("target_metrics", ""),
        p.get("target_partners", ""), p.get("offer", ""), p.get("timeline", ""),
        p.get("resources_required", ""), p.get("priority", ""), p.get("pic", ""),
        p.get("end_results", ""), p.get("note", ""),
    ] for i, p in enumerate(data.get("partnership_goals", []) or [{}])]
    table(ws, 3,
          ["No", "Goal Type", "Target Metrics", "Breakdown on target partners", "Offer",
           "Est Timeline", "Resources Required", "Priority", "PIC",
           "Update / end-results (write in past tense, before you start)", "Note"],
          rows, widths=[5, 20, 26, 30, 30, 14, 26, 10, 12, 44, 20])


SCHEMA = {
    "opportunity_ideas": [{"id": "Idea 1", "industry_segment": "", "problem_statement": "",
                           "target_audience": "", "technology_opportunity": "", "name": ""}],
    "decision_matrix": {"scale": "1-5",
                        "factors": [{"name": "Problem severity", "weight": 25}],
                        "ideas": [{"name": "", "scores": [0]}]},
    "company": {"name": "", "vision": "", "mission": "", "values": [], "culture": "",
                "corporate_objective": "", "mission_critical_priority": "", "milestones": []},
    "offer": {"problem": "", "solution": "", "existing_solutions": "",
              "differentiation_vs_existing": "", "why_no_existing": "", "differentiation": ""},
    "target_market": {"revenue_threshold": "", "revenue_model": "", "model_fit": "",
                      "why_buy": "", "willingness_to_pay": "", "ability_to_pay": ""},
    "buyer": {"pain_points": "", "motivation": "", "enables": "", "initiative": "",
              "is_decision_maker": "", "dm_influencers": "", "decision_maker": "",
              "route_to_dm": "", "is_purse_holder": "", "budget_requirements": "",
              "purse_holder": "", "route_to_budget": ""},
    "personas": [{"tier": "A - Must Win",
                  "company": {"industry": "", "revenue": "", "employee_size": "", "countries": ""},
                  "persona": {"job_title": "", "work_experience": "", "education": ""},
                  "relevance": "",
                  "attention": {"interaction_preferences": "", "content_types": "", "content_sources": ""}}],
    "methodologies": {"qualification": {"framework": "PNUB", "definitions": {"P": "", "N": "", "U": "", "B": ""}},
                      "philosophy": {"name": "Value Selling", "notes": "", "stage_gate": ""}},
    "sales_cycle": [{"stage": "", "entry_evidence": "", "goal": "", "exit_criteria": "",
                     "conversion_target": "", "median_days_target": ""}],
    "cold_outreach": {"cold_email": {"trigger": "", "bridge": "", "value_hypothesis": "",
                                     "cta": "", "subject": "", "body": ""},
                      "cold_call_pitch": "", "cold_call_question": "",
                      "touches": [{"n": 1, "day": 1, "channel": "Email", "content": ""}]},
    "objections": [{"objection": "", "underlying_cause": "", "exploring_question": "",
                    "response": "", "proof_asset": ""}],
    "partnership_goals": [{"no": 1, "goal_type": "", "target_metrics": "", "target_partners": "",
                           "offer": "", "timeline": "", "resources_required": "", "priority": "",
                           "pic": "", "end_results": "", "note": ""}],
}

BUILDERS = [
    ("opportunity_ideas", tab_brainstorm), ("decision_matrix", tab_decision_matrix),
    ("company", tab_mission), ("offer", tab_icp), ("personas", tab_personas),
    ("methodologies", tab_methodologies), ("sales_cycle", tab_cycle),
    ("cold_outreach", tab_outreach), ("objections", tab_objections),
    ("partnership_goals", tab_partnerships),
]


def main():
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("answers", nargs="?", help="Path to answers JSON")
    ap.add_argument("-o", "--output", default="GTM_filled.xlsx")
    ap.add_argument("--schema", action="store_true", help="Print the expected JSON shape and exit")
    ap.add_argument("--partnership-only", action="store_true", help="Emit only the Partnership Goals tab")
    ap.add_argument("--ideation-only", action="store_true",
                    help="Emit only the Opportunity Brainstorm and Decision Matrix tabs")
    args = ap.parse_args()

    if args.schema:
        print(json.dumps(SCHEMA, indent=2))
        return 0
    if not args.answers:
        ap.error("answers JSON required (or use --schema)")
    if not HAVE_OPENPYXL:
        sys.exit("build_gtm_workbook.py needs the 'openpyxl' package to write xlsx.\n"
                 "Install it with: pip install openpyxl")

    with open(args.answers) as f:
        data = json.load(f)

    wb = Workbook()
    wb.remove(wb.active)
    only = None
    if args.partnership_only:
        only = {"partnership_goals"}
    elif args.ideation_only:
        only = {"opportunity_ideas", "decision_matrix"}
    builders = [b for b in BUILDERS if only is None or b[0] in only]
    for _, build in builders:
        build(wb, data)
    wb.save(args.output)

    keys = [k for k, _ in builders] if only else list(SCHEMA)
    filled = [k for k in keys if data.get(k)]
    missing = [k for k in keys if not data.get(k)]
    print(json.dumps({"output": args.output, "filled_sections": filled,
                      "left_blank_tbd": missing}, indent=2))
    return 0


if __name__ == "__main__":
    sys.exit(main())
